"""
Data loader for the CCM GDP Projection Dashboard.
Reads all sheets from the Excel model into clean DataFrames.
"""
import pandas as pd
import numpy as np
import openpyxl
from pathlib import Path

EXCEL_PATH = Path(__file__).parent.parent / "model" / "Version 6.xlsx"

REGIONS = {
    "ASEAN": ["IDN", "MYS", "PHL", "SGP", "THA", "VNM", "MMR", "LAO", "KHM", "BRN"],
    "East Asia": ["CHN", "HKG", "JPN", "KOR", "MNG", "TWN"],
    "South Asia": ["BGD", "IND", "PAK", "LKA", "NPL", "BTN", "MDV"],
    "North America": ["USA", "CAN", "MEX"],
    "Latin America": [
        "ARG", "BRA", "CHL", "COL", "ECU", "PER", "VEN", "BOL", "PRY", "URY",
        "CRI", "GTM", "HND", "SLV", "NIC", "PAN", "DOM", "JAM", "HTI", "TTO",
        "GUY", "SUR",
    ],
    "Europe": [
        "DEU", "FRA", "GBR", "ITA", "ESP", "NLD", "BEL", "CHE", "SWE", "NOR",
        "DNK", "FIN", "POL", "CZE", "HUN", "ROU", "BGR", "GRC", "PRT", "AUT",
        "SVK", "SVN", "HRV", "LTU", "LVA", "EST", "IRL", "LUX", "RUS", "UKR",
        "BLR", "ALB", "BIH", "MKD", "SRB", "MNE", "ISL", "MLT", "CYP", "MDA",
    ],
    "Middle East & N. Africa": [
        "ARE", "SAU", "IRN", "IRQ", "EGY", "DZA", "MAR", "TUN", "JOR", "OMN",
        "KWT", "QAT", "BHR", "ISR", "LBN", "YEM",
    ],
    "Sub-Saharan Africa": [
        "NGA", "ZAF", "ETH", "KEN", "GHA", "TZA", "UGA", "CIV", "CMR", "SEN",
        "ZMB", "MDG", "MOZ", "AGO", "SDN", "MLI", "BFA", "GIN", "NER", "TCD",
        "ZWE", "RWA", "BEN", "MWI", "LSO", "BDI", "SLE", "LBR", "BWA", "NAM",
        "SWZ", "COD", "GAB", "GNQ", "MRT",
    ],
    "Oceania": ["AUS", "NZL", "PNG", "FJI"],
    "Central Asia & Caucasus": ["KAZ", "UZB", "TKM", "KGZ", "TJK", "GEO", "ARM", "AZE"],
}

G7 = ["USA", "CAN", "JPN", "DEU", "FRA", "GBR", "ITA"]

# Reverse map: code → region
CODE_TO_REGION = {}
for region, codes in REGIONS.items():
    for c in codes:
        CODE_TO_REGION[c] = region


def _load_time_series(ws, year_row: int, data_start_row: int, code_col: int = 1) -> pd.DataFrame:
    """Load a standard time-series sheet: rows=countries, cols=years."""
    year_to_col: dict[int, int] = {}
    for c in range(3, ws.max_column + 1):
        v = ws.cell(year_row, c).value
        if isinstance(v, (int, float)) and 1970 < v < 2110:
            year_to_col[int(v)] = c

    records: dict[str, dict] = {}
    for r in range(data_start_row, ws.max_row + 1):
        code = ws.cell(r, code_col).value
        if not isinstance(code, str) or len(code) < 2 or len(code) > 4:
            continue
        row: dict[int, float] = {}
        for yr, col in year_to_col.items():
            v = ws.cell(r, col).value
            if isinstance(v, (int, float)):
                row[yr] = float(v)
        if row and code not in records:   # keep first occurrence only
            records[code] = row

    df = pd.DataFrame(records).T
    df.index.name = "code"
    df.columns = df.columns.astype(int)
    return df


def load_data() -> dict:
    """Load all relevant sheets from the Excel model."""
    wb = openpyxl.load_workbook(str(EXCEL_PATH), data_only=True)

    # ── Core time-series ─────────────────────────────────────────────────────
    gdppc_df = _load_time_series(wb["gdppc"], year_row=2, data_start_row=3)
    pop_df   = _load_time_series(wb["population"], year_row=1, data_start_row=2)
    wap_df   = _load_time_series(wb["wap"],   year_row=1, data_start_row=2)
    rel_prod_df = _load_time_series(wb["Relative Productivity"], year_row=3, data_start_row=4)

    # Compute GDP (billions 2021 USD PPP) = GDPPC * Population(thousands) / 1,000,000
    common_codes = gdppc_df.index.intersection(pop_df.index)
    common_years = gdppc_df.columns.intersection(pop_df.columns)
    gdp_df = gdppc_df.loc[common_codes, common_years].multiply(
        pop_df.loc[common_codes, common_years]
    ) / 1_000_000

    # ── Kernel Summary ────────────────────────────────────────────────────────
    ws_ks = wb["Kernel Summary"]
    ks_rows = []
    for r in range(2, ws_ks.max_row + 1):
        name    = ws_ks.cell(r, 1).value
        code    = ws_ks.cell(r, 2).value
        gci     = ws_ks.cell(r, 3).value
        comp    = ws_ks.cell(r, 4).value
        lss_log = ws_ks.cell(r, 5).value
        lss_pct = ws_ks.cell(r, 6).value
        if isinstance(name, str) and isinstance(code, str):
            ks_rows.append({
                "country": name,
                "code": code,
                "gci": float(gci) if isinstance(gci, (int, float)) else None,
                "comparator": int(comp) if isinstance(comp, (int, float)) else 0,
                "lss_log": float(lss_log) if isinstance(lss_log, (int, float)) else None,
                "lss_pct": float(lss_pct) if isinstance(lss_pct, (int, float)) else None,
            })
    kernel_df = pd.DataFrame(ks_rows).set_index("code")

    # ── Comparators (for kernel re-estimation) ────────────────────────────────
    ws_c = wb["Comparators"]
    comp_rows = []
    for r in range(2, ws_c.max_row + 1):
        row = [ws_c.cell(r, c).value for c in range(1, 6)]
        if isinstance(row[0], str) and isinstance(row[1], str):
            comp_rows.append({
                "country":  row[0],
                "code":     row[1],
                "gci":      float(row[2]) if isinstance(row[2], (int, float)) else None,
                "log_prod": float(row[3]) if isinstance(row[3], (int, float)) else None,
                "comparator": int(row[4]) if isinstance(row[4], (int, float)) else 0,
            })
    comp_df = pd.DataFrame(comp_rows).dropna(subset=["gci", "log_prod"])

    # ── Default parameters ───────────────────────────────────────────────────
    ws_pc = wb["Parameters & Charts"]
    params = {
        "beta":         _safe_float(ws_pc.cell(9, 2).value, 0.025),
        "g_usa":        _safe_float(ws_pc.cell(9, 3).value, 0.018),
        "bandwidth":    _safe_float(ws_pc.cell(3, 7).value, 7.76),
        "gci_weight_a": _safe_float(ws_pc.cell(6, 3).value, 0.4),
        "gci_weight_b": _safe_float(ws_pc.cell(6, 4).value, 0.5),
        "gci_weight_c": _safe_float(ws_pc.cell(6, 5).value, 0.1),
    }

    # ── Country name map ─────────────────────────────────────────────────────
    name_map: dict[str, str] = kernel_df["country"].to_dict()
    for _, row in comp_df.iterrows():
        if row["code"] not in name_map:
            name_map[row["code"]] = row["country"]

    return {
        "gdppc":    gdppc_df,
        "pop":      pop_df,
        "wap":      wap_df,
        "rel_prod": rel_prod_df,
        "gdp":      gdp_df,
        "kernel":   kernel_df,
        "comparators": comp_df,
        "params":   params,
        "name_map": name_map,
    }


def _safe_float(v, default: float) -> float:
    return float(v) if isinstance(v, (int, float)) else default


def get_country_region(code: str) -> str:
    return CODE_TO_REGION.get(code, "Other")
