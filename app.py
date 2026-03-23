"""
CCM GDP Projection Dashboard
Interactive Streamlit app for the Conditional Convergence Model.
"""
import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
import plotly.express as px
from plotly.subplots import make_subplots

from utils.data_loader import load_data, REGIONS, G7, get_country_region
from utils.model import (
    project_country,
    reestimate_lss,
    kernel_curve,
    regional_gdp,
)

# ─── Page config ─────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="CCM GDP Projection Dashboard",
    page_icon=None,
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Open+Sans:wght@400;600;700&display=swap');

    /* Base */
    html, body, [class*="css"] {
        font-family: 'Open Sans', Arial, sans-serif !important;
        border-radius: 0px !important;
    }
    *, *::before, *::after {
        border-radius: 0px !important;
    }

    /* Header bar */
    header[data-testid="stHeader"] {
        background-color: #ffffff;
        border-bottom: 2px solid #0073CF;
    }

    /* Sidebar */
    section[data-testid="stSidebar"] {
        background-color: #f7f9fc;
        border-right: 1px solid #dce3ea;
    }
    section[data-testid="stSidebar"] h2 {
        font-size: 0.78rem;
        font-weight: 700;
        text-transform: uppercase;
        letter-spacing: 0.08em;
        color: #4a6785;
        margin-top: 1rem;
        margin-bottom: 0.25rem;
    }

    /* Main content */
    .main .block-container {
        padding-top: 2rem;
        max-width: 1200px;
    }

    /* Tabs */
    button[data-baseweb="tab"] {
        font-size: 0.82rem;
        font-weight: 600;
        text-transform: uppercase;
        letter-spacing: 0.05em;
        color: #4a6785 !important;
        border-bottom: 2px solid transparent;
    }
    button[data-baseweb="tab"][aria-selected="true"] {
        color: #0073CF !important;
        border-bottom: 2px solid #0073CF;
    }

    /* Subheaders */
    h2, h3 {
        color: #1a3a5c !important;
        font-weight: 700 !important;
        letter-spacing: -0.01em;
    }
    h4, h5 {
        color: #2c5282 !important;
        font-weight: 600 !important;
    }

    /* Metric / info boxes */
    [data-testid="metric-container"] {
        background-color: #f0f4f9;
        border-left: 3px solid #0073CF;
        padding: 0.75rem 1rem;
    }

    /* Buttons */
    .stButton > button {
        background-color: #0073CF;
        color: white;
        font-weight: 600;
        font-size: 0.82rem;
        text-transform: uppercase;
        letter-spacing: 0.04em;
        border: none;
        padding: 0.5rem 1.25rem;
    }
    .stButton > button:hover {
        background-color: #005fa3;
    }

    /* Dataframe / table */
    [data-testid="stDataFrame"] {
        border: 1px solid #dce3ea;
    }

    /* Dividers */
    hr {
        border-color: #dce3ea;
        margin: 1rem 0;
    }

    /* Expanders */
    details summary {
        font-size: 0.82rem;
        font-weight: 600;
        text-transform: uppercase;
        letter-spacing: 0.05em;
        color: #4a6785;
    }

    /* Caption text */
    [data-testid="stCaptionContainer"] p {
        color: #6b7c93;
        font-size: 0.78rem;
    }

    /* Input labels */
    label[data-testid="stWidgetLabel"] p {
        font-size: 0.78rem;
        font-weight: 600;
        color: #2c5282;
        text-transform: uppercase;
        letter-spacing: 0.04em;
    }
</style>
""", unsafe_allow_html=True)

COLORS = px.colors.qualitative.Plotly
HIST_COLOR = "rgba(100,100,100,0.3)"

# ─── Load data (cached) ───────────────────────────────────────────────────────
@st.cache_data(show_spinner="Loading model data…")
def get_data():
    return load_data()

data      = get_data()
gdppc_df  = data["gdppc"]
pop_df    = data["pop"]
wap_df    = data["wap"]
rel_prod_df = data["rel_prod"]
gdp_df    = data["gdp"]
kernel_df = data["kernel"]
comp_df   = data["comparators"]
name_map  = data["name_map"]
defaults  = data["params"]

HIST_END   = 2024
PROJ_START = HIST_END

all_codes = sorted(
    set(gdppc_df.index) & set(kernel_df.index) & set(rel_prod_df.index)
)

def label(code: str) -> str:
    return f"{name_map.get(code, code)} ({code})"

# ─── Session state defaults ───────────────────────────────────────────────────
if "beta_key" not in st.session_state:
    st.session_state["beta_key"] = float(defaults["beta"])
if "g_usa_key" not in st.session_state:
    st.session_state["g_usa_key"] = float(defaults["g_usa"]) * 100

# ─── Sidebar ─────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("## Model Controls")

    st.markdown("## Scenario Presets")
    _sc1, _sc2, _sc3 = st.columns(3)
    if _sc1.button("Baseline", use_container_width=True):
        st.session_state["beta_key"] = float(defaults["beta"])
        st.session_state["g_usa_key"] = float(defaults["g_usa"]) * 100
    if _sc2.button("High Growth", use_container_width=True):
        st.session_state["beta_key"] = 0.04
        st.session_state["g_usa_key"] = 2.5
    if _sc3.button("Stagnation", use_container_width=True):
        st.session_state["beta_key"] = 0.01
        st.session_state["g_usa_key"] = 1.0

    with st.expander("Convergence Parameters", expanded=True):
        beta = st.slider(
            "β — Convergence Speed (annual)",
            0.001, 0.10, float(defaults["beta"]), 0.001,
            format="%.3f",
            key="beta_key",
            help="Rate at which each country's productivity gap closes each year. "
                 "Original model: 0.025 (2.5 % per year).",
        )
        g_usa = st.slider(
            "g_USA — US Productivity Growth (%/yr)",
            0.0, 5.0, float(defaults["g_usa"]) * 100, 0.1,
            format="%.1f%%",
            key="g_usa_key",
            help="Annual growth rate of the US productivity frontier.",
        ) / 100
        show_bands = st.checkbox("Show uncertainty bands", value=False,
            help="Shaded range between β×0.6 and β×1.4 projections.")

    with st.expander("Kernel Estimation", expanded=False):
        bandwidth = st.slider(
            "Kernel Bandwidth",
            1.0, 25.0, float(defaults["bandwidth"]), 0.25,
            help="Smoothing parameter for Gaussian kernel regression "
                 "(GCI → Long-run Steady State). Original: 7.76.",
        )
        st.caption(
            "Higher bandwidth → smoother LSS curve, less differentiation between countries."
        )

    with st.expander("GCI Weights", expanded=False):
        st.caption(
            "World Economic Forum weights for the three GCI pillars. "
            "These determine which economy-type model governs each country's steady state."
        )
        col_a, col_b, col_c = st.columns(3)
        with col_a:
            w_a = st.number_input("A: Basic Req.", 0.0, 1.0, float(defaults["gci_weight_a"]), 0.05, format="%.2f")
        with col_b:
            w_b = st.number_input("B: Efficiency", 0.0, 1.0, float(defaults["gci_weight_b"]), 0.05, format="%.2f")
        with col_c:
            w_c = st.number_input("C: Innovation", 0.0, 1.0, float(defaults["gci_weight_c"]), 0.05, format="%.2f")
        total_w = w_a + w_b + w_c
        if abs(total_w - 1.0) > 0.015:
            st.warning(f"Weights sum to {total_w:.2f} (should be 1.0)")
        econ_types = {
            "Factor-driven (0.60 / 0.35 / 0.05)":       (0.60, 0.35, 0.05),
            "Efficiency-driven (0.40 / 0.50 / 0.10)":   (0.40, 0.50, 0.10),
            "Innovation-driven (0.20 / 0.50 / 0.30)":   (0.20, 0.50, 0.30),
        }
        closest = min(econ_types, key=lambda k: sum((a-b)**2 for a, b in zip(econ_types[k], (w_a, w_b, w_c))))
        st.info(f"Closest preset: **{closest.split('(')[0].strip()}**")

    st.divider()
    st.markdown("## Country / Region")

    default_primary = "CHN" if "CHN" in all_codes else all_codes[0]
    primary = st.selectbox(
        "Primary Country",
        all_codes,
        index=all_codes.index(default_primary),
        format_func=label,
    )
    compare_with = st.multiselect(
        "Compare With",
        [c for c in all_codes if c != primary],
        default=[c for c in ["USA", "IND", "JPN", "IDN"] if c in all_codes and c != primary],
        format_func=label,
    )
    selected = [primary] + compare_with

    st.divider()
    st.markdown("## Projection")
    proj_end = st.slider("End Year", 2030, 2050, 2050, 1)
    show_original = st.checkbox("Overlay original Excel projection", value=True)

# ─── Re-estimate LSS with current bandwidth ───────────────────────────────────
@st.cache_data(show_spinner=False)
def compute_lss(bw: float) -> dict:
    return reestimate_lss(kernel_df, comp_df, bw)

lss_map = compute_lss(bandwidth)

# ─── Run projections ──────────────────────────────────────────────────────────
@st.cache_data(show_spinner=False)
def run_projections(codes: list, beta_: float, g_usa_: float, bw: float, end: int) -> dict:
    lss = compute_lss(bw)
    results = {}
    for code in codes:
        lss_val = lss.get(code)
        if lss_val is None or not np.isfinite(lss_val):
            continue
        proj = project_country(
            code, gdppc_df, pop_df, wap_df, rel_prod_df,
            lss_val, beta_, g_usa_,
            base_year=PROJ_START, end_year=end,
        )
        if proj is not None:
            results[code] = proj
    return results

projections = run_projections(tuple(selected), beta, g_usa, bandwidth, proj_end)

# ─── Uncertainty band projections (β ±40%) ────────────────────────────────────
beta_lo = max(0.001, beta * 0.6)
beta_hi = min(0.10,  beta * 1.4)
proj_lo = run_projections(tuple(selected), beta_lo, g_usa, bandwidth, proj_end)
proj_hi = run_projections(tuple(selected), beta_hi, g_usa, bandwidth, proj_end)

# ─── Helper: build combined historical + projection series ────────────────────
def build_series(code: str, metric: str) -> tuple[pd.Series, pd.Series, pd.Series | None]:
    """Return (historical, projection_custom, projection_original)."""
    src_df = gdppc_df if metric == "gdppc" else gdp_df
    hist_years = [y for y in range(1980, HIST_END + 1) if y in src_df.columns]
    hist = src_df.loc[code, hist_years].dropna() if code in src_df.index else pd.Series(dtype=float)

    proj_col = "gdppc" if metric == "gdppc" else "gdp_bn"
    proj_custom = projections[code][proj_col] if code in projections else pd.Series(dtype=float)

    # Original Excel projection (already stored in gdppc_df / gdp_df for 2025+)
    orig = None
    if show_original and metric == "gdppc" and code in gdppc_df.index:
        orig_years = [y for y in range(HIST_END, proj_end + 1) if y in gdppc_df.columns]
        orig = gdppc_df.loc[code, orig_years].dropna()

    return hist, proj_custom, orig


def fmt_gdp_trn(v: float) -> str:
    """Format a value already in Trillions."""
    return f"${v:.2f}T"


def fmt_gdppc(v: float) -> str:
    return f"${v:,.0f}"


# ─── Build chart helpers ──────────────────────────────────────────────────────
def line_chart(title: str, metric: str, unit_label: str, fmt_fn, divisor: float = 1.0,
               annotate_crossovers: bool = False) -> go.Figure:
    fig = go.Figure()
    proj_col = "gdppc" if metric == "gdppc" else "gdp_bn"

    for i, code in enumerate(selected):
        color = COLORS[i % len(COLORS)]
        hist, proj, orig = build_series(code, metric)
        cname = name_map.get(code, code)

        if not hist.empty:
            yvals = hist.values / divisor
            fig.add_trace(go.Scatter(
                x=list(hist.index), y=list(yvals),
                name=cname,
                line=dict(color=color, width=2),
                mode="lines",
                legendgroup=code,
                hovertemplate=f"<b>{cname}</b><br>%{{x}}: %{{customdata}}<extra></extra>",
                customdata=[fmt_fn(v) for v in yvals],
            ))

        # Uncertainty band
        if show_bands and code in proj_lo and code in proj_hi:
            _lo = proj_lo[code][proj_col]
            _hi = proj_hi[code][proj_col]
            common_idx = _lo.index.intersection(_hi.index)
            if len(common_idx) > 0:
                import plotly.colors as pc
                rgb = pc.hex_to_rgb(color) if color.startswith("#") else (100, 100, 200)
                band_color = f"rgba({rgb[0]},{rgb[1]},{rgb[2]},0.12)"
                fig.add_trace(go.Scatter(
                    x=list(common_idx) + list(common_idx[::-1]),
                    y=list(_hi[common_idx].values / divisor) + list(_lo[common_idx].values / divisor),
                    fill="toself", fillcolor=band_color,
                    line=dict(width=0), showlegend=False, legendgroup=code,
                    hoverinfo="skip",
                ))

        if not proj.empty:
            yvals = proj.values / divisor
            fig.add_trace(go.Scatter(
                x=list(proj.index), y=list(yvals),
                name=f"{cname} (model)",
                line=dict(color=color, width=2, dash="dash"),
                mode="lines",
                legendgroup=code,
                showlegend=False,
                hovertemplate=f"<b>{cname} (model)</b><br>%{{x}}: %{{customdata}}<extra></extra>",
                customdata=[fmt_fn(v) for v in yvals],
            ))
        if orig is not None and not orig.empty:
            yvals = orig.values / divisor
            fig.add_trace(go.Scatter(
                x=list(orig.index), y=list(yvals),
                name=f"{cname} (original)",
                line=dict(color=color, width=1.5, dash="dot"),
                mode="lines",
                legendgroup=code,
                showlegend=False,
                opacity=0.6,
                hovertemplate=f"<b>{cname} (original)</b><br>%{{x}}: %{{customdata}}<extra></extra>",
                customdata=[fmt_fn(v) for v in yvals],
            ))

    # Crossover annotations
    if annotate_crossovers and len(selected) >= 2:
        for i in range(len(selected)):
            for j in range(i + 1, len(selected)):
                ca, cb = selected[i], selected[j]
                if ca not in projections or cb not in projections:
                    continue
                sa = projections[ca][proj_col]
                sb = projections[cb][proj_col]
                common = sa.index.intersection(sb.index)
                if len(common) < 2:
                    continue
                diff = sa[common] - sb[common]
                for k in range(1, len(diff)):
                    if diff.iloc[k - 1] * diff.iloc[k] < 0:
                        cx_yr = int(diff.index[k])
                        ca_name = name_map.get(ca, ca)
                        cb_name = name_map.get(cb, cb)
                        label_txt = f"{ca_name} / {cb_name} cross"
                        fig.add_vline(
                            x=cx_yr, line_dash="dot", line_color="#888", line_width=1,
                            annotation_text=label_txt,
                            annotation_position="top",
                            annotation_font_size=10,
                            annotation_font_color="#555",
                        )
                        break

    # Shaded historical region
    fig.add_vrect(x0=1980, x1=HIST_END, fillcolor="gray", opacity=0.06, line_width=0)
    fig.add_vline(x=HIST_END, line_dash="dot", line_color="gray", line_width=1)

    fig.update_layout(
        title=title, height=520,
        xaxis_title="Year", yaxis_title=unit_label,
        hovermode="x unified",
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
        plot_bgcolor="white",
        paper_bgcolor="white",
        xaxis=dict(showgrid=True, gridcolor="#eee"),
        yaxis=dict(showgrid=True, gridcolor="#eee"),
        margin=dict(t=60, b=40),
    )
    return fig


# ─── Summary panel ────────────────────────────────────────────────────────────
_prim = primary
_prim_name = name_map.get(_prim, _prim)
_sm_cols = st.columns(4)

with _sm_cols[0]:
    if _prim in projections and proj_end in projections[_prim].index:
        _gdp_val = float(projections[_prim].loc[proj_end, "gdp_bn"]) / 1000
        st.metric(f"{_prim_name} GDP ({proj_end})", f"${_gdp_val:.1f}T")
    else:
        st.metric(f"{_prim_name} GDP ({proj_end})", "—")

with _sm_cols[1]:
    if _prim in projections and proj_end in projections[_prim].index:
        _gdppc_val = float(projections[_prim].loc[proj_end, "gdppc"])
        st.metric(f"{_prim_name} GDP/capita ({proj_end})", f"${_gdppc_val:,.0f}")
    else:
        st.metric(f"{_prim_name} GDP/capita ({proj_end})", "—")

with _sm_cols[2]:
    _ranked = sorted(
        [(c, float(projections[c].loc[proj_end, "gdp_bn"])) for c in selected
         if c in projections and proj_end in projections[c].index],
        key=lambda x: -x[1]
    )
    if _ranked:
        _top_code, _top_val = _ranked[0]
        st.metric(f"Largest Economy ({proj_end})", name_map.get(_top_code, _top_code), f"${_top_val/1000:.1f}T")
    else:
        st.metric(f"Largest Economy ({proj_end})", "—")

with _sm_cols[3]:
    _crossover_yr = None
    if _prim != "USA" and _prim in projections and "USA" in projections:
        for _yr in range(HIST_END + 1, proj_end + 1):
            if _yr in projections[_prim].index and _yr in projections["USA"].index:
                _v_prim = float(projections[_prim].loc[_yr, "gdp_bn"])
                _v_usa  = float(projections["USA"].loc[_yr, "gdp_bn"])
                if _v_prim >= _v_usa:
                    _crossover_yr = _yr
                    break
    if _crossover_yr:
        st.metric(f"{_prim_name} overtakes USA", str(_crossover_yr))
    elif _prim == "USA":
        st.metric("USA", "Frontier economy")
    else:
        st.metric(f"{_prim_name} vs USA", "No overtake by 2050")

st.divider()

# ─── TABS ─────────────────────────────────────────────────────────────────────
tab_gdp, tab_gdppc, tab_region, tab_conv, tab_ppt, tab_tech = st.tabs([
    "GDP (Total)",
    "GDP per Capita",
    "Regional View",
    "Convergence",
    "Presentation Charts",
    "Technical",
])

# ════════════════════════════════════════════════════════════════════════════════
# TAB 1 — GDP Total
# ════════════════════════════════════════════════════════════════════════════════
with tab_gdp:
    st.subheader("Total GDP — 2021 USD PPP (Trillions)")
    st.caption(
        "Solid line = historical data (1980–2024) | Dashed = model projection | "
        "Dotted = original Excel baseline"
    )

    fig = line_chart(
        "GDP (Trillions, 2021 USD PPP)", "gdp",
        "Trillions (2021 USD PPP)", fmt_gdp_trn,
        divisor=1000, annotate_crossovers=True,
    )
    st.plotly_chart(fig, use_container_width=True,
        config={"toImageButtonOptions": {"format": "png", "filename": "gdp_total", "scale": 2}})

    # Summary table
    st.markdown("#### 2024 → Projection Snapshot")
    rows = []
    for code in selected:
        if code not in projections:
            continue
        proj = projections[code]
        snap_years = [y for y in [2030, 2040, 2050, proj_end] if y <= proj_end]
        snap_years = sorted(set(snap_years))
        row = {"Country": name_map.get(code, code), "Code": code,
               "GDP 2024 (T)": fmt_gdp_trn(float(gdp_df.loc[code, 2024]) / 1000) if code in gdp_df.index else "—"}
        for yr in snap_years:
            if yr in proj.index:
                row[f"GDP {yr} (T)"] = fmt_gdp_trn(float(proj.loc[yr, "gdp_bn"]) / 1000)
        rows.append(row)
    if rows:
        st.dataframe(pd.DataFrame(rows).set_index("Code"), use_container_width=True)

    # 2050 GDP ranking
    _rank_rows = []
    for code in selected:
        if code not in projections:
            continue
        _p = projections[code]
        if proj_end in _p.index:
            _rank_rows.append({
                "Country": name_map.get(code, code),
                "Code": code,
                f"GDP {proj_end} (T)": round(float(_p.loc[proj_end, "gdp_bn"]) / 1000, 2),
                f"GDP/capita {proj_end}": int(float(_p.loc[proj_end, "gdppc"])),
            })
    if _rank_rows:
        st.markdown(f"#### {proj_end} Rankings — Selected Countries")
        _rank_df = pd.DataFrame(_rank_rows).sort_values(f"GDP {proj_end} (T)", ascending=False).reset_index(drop=True)
        _rank_df.index += 1
        _rank_df[f"GDP/capita {proj_end}"] = _rank_df[f"GDP/capita {proj_end}"].apply(lambda x: f"${x:,}")
        _rank_df[f"GDP {proj_end} (T)"] = _rank_df[f"GDP {proj_end} (T)"].apply(lambda x: f"${x:.2f}T")
        st.dataframe(_rank_df.set_index("Code"), use_container_width=True)


# ════════════════════════════════════════════════════════════════════════════════
# TAB 2 — GDP per Capita
# ════════════════════════════════════════════════════════════════════════════════
with tab_gdppc:
    st.subheader("GDP per Capita — 2021 USD PPP")
    st.caption(
        "Solid line = historical (1980–2024) | Dashed = model projection | "
        "Dotted = original Excel baseline"
    )

    fig2 = line_chart("GDP per Capita (2021 USD PPP)", "gdppc", "USD", fmt_gdppc)
    st.plotly_chart(fig2, use_container_width=True,
        config={"toImageButtonOptions": {"format": "png", "filename": "gdp_per_capita", "scale": 2}})

    # GDPPC relative to USA
    st.markdown("#### GDP per Capita as % of USA")
    fig_rel = go.Figure()
    usa_gdppc = gdppc_df.loc["USA"] if "USA" in gdppc_df.index else None
    for i, code in enumerate(selected):
        if code == "USA" or usa_gdppc is None:
            continue
        color = COLORS[i % len(COLORS)]
        cname = name_map.get(code, code)
        hist_yrs = [y for y in range(1980, HIST_END + 1) if y in gdppc_df.columns]
        if code in gdppc_df.index:
            hist_rel = (gdppc_df.loc[code, hist_yrs] / usa_gdppc[hist_yrs] * 100).dropna()
            fig_rel.add_trace(go.Scatter(
                x=list(hist_rel.index), y=list(hist_rel.values),
                name=cname, line=dict(color=color, width=2), legendgroup=code,
                hovertemplate=f"<b>{cname}</b><br>%{{x}}: %{{y:.1f}}% of USA<extra></extra>",
            ))
        if code in projections:
            proj = projections[code]["gdppc"]
            usa_proj = [
                float(gdppc_df.loc["USA", y]) if y in gdppc_df.columns
                else float(gdppc_df.loc["USA", HIST_END]) * np.exp(g_usa * (y - HIST_END))
                for y in proj.index
            ]
            proj_rel = proj.values / np.array(usa_proj) * 100
            fig_rel.add_trace(go.Scatter(
                x=list(proj.index), y=list(proj_rel),
                line=dict(color=color, width=2, dash="dash"), showlegend=False, legendgroup=code,
                hovertemplate=f"<b>{cname} (proj)</b><br>%{{x}}: %{{y:.1f}}% of USA<extra></extra>",
            ))
    fig_rel.add_vrect(x0=1980, x1=HIST_END, fillcolor="gray", opacity=0.06, line_width=0)
    fig_rel.add_vline(x=HIST_END, line_dash="dot", line_color="gray", line_width=1)
    fig_rel.add_hline(y=100, line_dash="dash", line_color="black", line_width=1, annotation_text="USA = 100%")
    fig_rel.update_layout(
        height=450, yaxis_title="% of USA GDP per Capita",
        xaxis_title="Year", hovermode="x unified",
        plot_bgcolor="white", paper_bgcolor="white",
        xaxis=dict(showgrid=True, gridcolor="#eee"),
        yaxis=dict(showgrid=True, gridcolor="#eee"),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
        margin=dict(t=40, b=40),
    )
    st.plotly_chart(fig_rel, use_container_width=True)


# ════════════════════════════════════════════════════════════════════════════════
# TAB 3 — Regional View
# ════════════════════════════════════════════════════════════════════════════════
with tab_region:
    st.subheader("Regional GDP — 2021 USD PPP (Trillions)")

    col_r1, col_r2 = st.columns([2, 1])
    with col_r2:
        selected_regions = st.multiselect(
            "Regions to show",
            list(REGIONS.keys()),
            default=list(REGIONS.keys()),
        )
        chart_type = st.radio("Chart type", ["Stacked Area", "Line", "Share of World (%)"], index=0)

    all_years = list(range(1980, proj_end + 1))
    # For historical years, use Excel data; for projection years, sum individual country Excel projections
    region_gdp_data: dict[str, pd.Series] = {}
    for region in selected_regions:
        codes = REGIONS[region]
        available = [c for c in codes if c in gdp_df.index]
        if not available:
            continue
        vals = {}
        for yr in all_years:
            if yr in gdp_df.columns:
                s = gdp_df.loc[available, yr].dropna().sum()
                vals[yr] = float(s)
        region_gdp_data[region] = pd.Series(vals)

    world_gdp = pd.Series({
        yr: float(gdp_df[yr].dropna().sum()) for yr in all_years if yr in gdp_df.columns
    })

    with col_r1:
        fig_reg = go.Figure()
        reg_colors = px.colors.qualitative.Set2
        for i, region in enumerate(selected_regions):
            if region not in region_gdp_data:
                continue
            series = region_gdp_data[region].dropna()
            c = reg_colors[i % len(reg_colors)]
            if chart_type == "Share of World (%)":
                yvals = (series / world_gdp.reindex(series.index) * 100).values
                ytitle = "Share of World GDP (%)"
            else:
                yvals = series.values / 1000  # → trillions
                ytitle = "GDP (Trillions, 2021 USD PPP)"

            if chart_type == "Stacked Area":
                fig_reg.add_trace(go.Scatter(
                    x=list(series.index), y=list(yvals),
                    name=region, fill="tonexty" if i > 0 else "tozeroy",
                    mode="lines", line=dict(color=c, width=0.5),
                    stackgroup="one",
                    hovertemplate=f"<b>{region}</b><br>%{{x}}: %{{y:.1f}}<extra></extra>",
                ))
            else:
                fig_reg.add_trace(go.Scatter(
                    x=list(series.index), y=list(yvals),
                    name=region, mode="lines", line=dict(color=c, width=2),
                    hovertemplate=f"<b>{region}</b><br>%{{x}}: %{{y:.1f}}<extra></extra>",
                ))

        fig_reg.add_vrect(x0=1980, x1=HIST_END, fillcolor="gray", opacity=0.05, line_width=0)
        fig_reg.add_vline(x=HIST_END, line_dash="dot", line_color="gray", line_width=1)
        fig_reg.update_layout(
            height=520, hovermode="x unified",
            xaxis_title="Year", yaxis_title=ytitle,
            plot_bgcolor="white", paper_bgcolor="white",
            xaxis=dict(showgrid=True, gridcolor="#eee"),
            yaxis=dict(showgrid=True, gridcolor="#eee"),
            legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
            margin=dict(t=40, b=40),
        )
        st.plotly_chart(fig_reg, use_container_width=True)


# ════════════════════════════════════════════════════════════════════════════════
# TAB 4 — Convergence Analysis
# ════════════════════════════════════════════════════════════════════════════════
with tab_conv:
    st.subheader("Productivity Convergence")

    col_c1, col_c2 = st.columns(2)

    # Left: convergence trajectory in log relative productivity
    with col_c1:
        st.markdown("##### Log Relative Productivity Trajectories")
        st.caption("Convergence of each country's labour productivity toward its long-run steady state (USA = 0)")

        fig_conv = go.Figure()
        for i, code in enumerate(selected):
            color = COLORS[i % len(COLORS)]
            cname = name_map.get(code, code)
            lss_val = lss_map.get(code)
            if lss_val is None:
                continue

            # Historical log relative productivity
            hist_rp_yrs = [y for y in range(1980, HIST_END + 1) if y in rel_prod_df.columns]
            if code in rel_prod_df.index:
                rp_hist = rel_prod_df.loc[code, hist_rp_yrs].dropna()
                log_rp_hist = np.log(rp_hist.clip(lower=1e-6))
                fig_conv.add_trace(go.Scatter(
                    x=list(log_rp_hist.index), y=list(log_rp_hist.values),
                    name=cname, line=dict(color=color, width=2), legendgroup=code,
                    hovertemplate=f"<b>{cname}</b><br>%{{x}}: log(rel. prod) = %{{y:.3f}}<extra></extra>",
                ))

            # Projected log relative productivity
            try:
                rp_2024 = float(rel_prod_df.loc[code, HIST_END])
                if rp_2024 > 0:
                    from utils.model import _convergence_trajectory
                    n = proj_end - HIST_END
                    traj = _convergence_trajectory(np.log(rp_2024), lss_val, beta, n)
                    proj_years = list(range(HIST_END, proj_end + 1))
                    fig_conv.add_trace(go.Scatter(
                        x=proj_years, y=list(traj),
                        line=dict(color=color, width=2, dash="dash"),
                        showlegend=False, legendgroup=code,
                        hovertemplate=f"<b>{cname} (proj)</b><br>%{{x}}: %{{y:.3f}}<extra></extra>",
                    ))
                    # LSS line
                    fig_conv.add_hline(
                        y=lss_val, line_dash="dot", line_color=color,
                        line_width=1, opacity=0.5,
                        annotation_text=f"{code} LSS",
                        annotation_position="right",
                    )
            except (KeyError, TypeError, ValueError):
                pass

        fig_conv.add_hline(y=0, line_dash="dash", line_color="black", line_width=1.5,
                           annotation_text="USA (frontier)", annotation_position="right")
        fig_conv.add_vrect(x0=1980, x1=HIST_END, fillcolor="gray", opacity=0.06, line_width=0)
        fig_conv.add_vline(x=HIST_END, line_dash="dot", line_color="gray", line_width=1)
        fig_conv.update_layout(
            height=460, hovermode="x unified",
            xaxis_title="Year",
            yaxis_title="ln(Productivity / USA Productivity)",
            plot_bgcolor="white", paper_bgcolor="white",
            xaxis=dict(showgrid=True, gridcolor="#eee"),
            yaxis=dict(showgrid=True, gridcolor="#eee"),
            legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
            margin=dict(t=30, b=40),
        )
        st.plotly_chart(fig_conv, use_container_width=True)

    # Right: current productivity gap vs LSS
    with col_c2:
        st.markdown("##### Current Gap vs Long-Run Steady State")
        st.caption(
            "Each point is a country. X = current log relative productivity (2023/24), "
            "Y = LSS. Points above the diagonal are overshooting (converging down)."
        )

        all_display = kernel_df.copy()
        all_display = all_display[all_display["lss_log"].notna()]

        curr_rp_col = max([y for y in rel_prod_df.columns if y <= HIST_END], default=None)
        scatter_data = []
        for code in all_display.index:
            if code in rel_prod_df.index and curr_rp_col:
                rp = rel_prod_df.loc[code, curr_rp_col]
                if isinstance(rp, (int, float)) and rp > 0:
                    scatter_data.append({
                        "code": code,
                        "country": name_map.get(code, code),
                        "log_rp": np.log(float(rp)),
                        "lss": float(lss_map.get(code, all_display.loc[code, "lss_log"])),
                        "region": get_country_region(code),
                        "selected": code in selected,
                    })
        sdf = pd.DataFrame(scatter_data)

        if not sdf.empty:
            fig_scatter = px.scatter(
                sdf, x="log_rp", y="lss",
                color="region", hover_name="country",
                hover_data={"code": True, "log_rp": ":.3f", "lss": ":.3f", "region": False},
                labels={"log_rp": "Current log(Rel. Productivity)", "lss": "Long-Run Steady State (LSS)"},
                height=460,
            )
            # Diagonal reference line
            lims = [min(sdf["log_rp"].min(), sdf["lss"].min()) - 0.1,
                    max(sdf["log_rp"].max(), sdf["lss"].max()) + 0.1]
            fig_scatter.add_trace(go.Scatter(
                x=lims, y=lims, mode="lines",
                line=dict(color="black", dash="dash", width=1),
                name="Diagonal (no gap)", showlegend=False,
            ))
            # Highlight selected countries
            sel_sdf = sdf[sdf["selected"]]
            if not sel_sdf.empty:
                fig_scatter.add_trace(go.Scatter(
                    x=sel_sdf["log_rp"], y=sel_sdf["lss"],
                    mode="markers+text", text=sel_sdf["code"],
                    textposition="top center",
                    marker=dict(size=12, color="black", symbol="star"),
                    name="Selected", showlegend=True,
                ))
            fig_scatter.update_layout(
                plot_bgcolor="white", paper_bgcolor="white",
                margin=dict(t=30, b=40),
            )
            st.plotly_chart(fig_scatter, use_container_width=True)


# ════════════════════════════════════════════════════════════════════════════════
# TAB 5 — PPT Charts
# ════════════════════════════════════════════════════════════════════════════════
with tab_ppt:
    st.subheader("Presentation Charts")
    st.caption(
        "Replication of key charts from the original presentation slides, "
        "extended to the selected projection horizon with model parameters."
    )

    ppt1, ppt2, ppt3, ppt4 = st.tabs([
        "Slide 12 — Regional GDP",
        "Slides 14–21 — Asia Focus",
        "Slide 25 — Forecast Comparison",
        "China vs G7 vs North America",
    ])

    # ── Slide 12: Asia share of world output ─────────────────────────────────
    with ppt1:
        st.markdown("##### Regional Share of World GDP (%, 1980–" + str(proj_end) + ")")
        ppt_regions_12 = [
            "ASEAN", "East Asia", "South Asia", "Europe",
            "North America", "Latin America",
            "Middle East & N. Africa", "Sub-Saharan Africa",
        ]
        ppt_years = list(range(1980, proj_end + 1))
        world_s = {yr: float(gdp_df[yr].dropna().sum()) for yr in ppt_years if yr in gdp_df.columns}

        fig_p12 = go.Figure()
        ppt_colors = px.colors.qualitative.Pastel
        for i, reg in enumerate(ppt_regions_12):
            codes_r = REGIONS.get(reg, [])
            avail = [c for c in codes_r if c in gdp_df.index]
            if not avail:
                continue
            vals = {}
            for yr in ppt_years:
                if yr in gdp_df.columns:
                    vals[yr] = float(gdp_df.loc[avail, yr].dropna().sum())
            s = pd.Series(vals)
            share = (s / pd.Series(world_s) * 100).dropna()
            fig_p12.add_trace(go.Scatter(
                x=list(share.index), y=list(share.values),
                name=reg, fill="tonexty" if i > 0 else "tozeroy",
                mode="lines", line=dict(width=0.5, color=ppt_colors[i % len(ppt_colors)]),
                stackgroup="one",
                hovertemplate=f"<b>{reg}</b><br>%{{x}}: %{{y:.1f}}%<extra></extra>",
            ))
        fig_p12.add_vrect(x0=1980, x1=HIST_END, fillcolor="gray", opacity=0.05, line_width=0)
        fig_p12.add_vline(x=HIST_END, line_dash="dot", line_color="gray", line_width=1)
        fig_p12.update_layout(
            height=520, hovermode="x unified",
            xaxis_title="Year", yaxis_title="Share of World GDP (%)",
            plot_bgcolor="white", paper_bgcolor="white",
            yaxis=dict(range=[0, 100], showgrid=True, gridcolor="#eee"),
            xaxis=dict(showgrid=True, gridcolor="#eee"),
            legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
            margin=dict(t=40, b=40),
        )
        st.plotly_chart(fig_p12, use_container_width=True)

    # ── Slides 14–21: Asia key economies ─────────────────────────────────────
    with ppt2:
        st.markdown("##### Key Asian Economies — GDP (Trillions, 2021 USD PPP)")
        asia_focus = ["CHN", "IND", "IDN", "JPN", "KOR", "VNM", "MYS", "SGP", "THA"]
        asia_available = [c for c in asia_focus if c in gdp_df.index]
        ppt_yr_range = list(range(1980, proj_end + 1))

        fig_asia = go.Figure()
        for i, code in enumerate(asia_available):
            cname = name_map.get(code, code)
            yr_avail = [y for y in ppt_yr_range if y in gdp_df.columns]
            s = gdp_df.loc[code, yr_avail].dropna()
            hist_s = s[s.index <= HIST_END]
            proj_s = s[s.index > HIST_END]
            c = COLORS[i % len(COLORS)]
            if not hist_s.empty:
                fig_asia.add_trace(go.Scatter(
                    x=list(hist_s.index), y=list(hist_s.values / 1000),
                    name=cname, line=dict(color=c, width=2), legendgroup=code,
                    hovertemplate=f"<b>{cname}</b><br>%{{x}}: $%{{y:.2f}}T<extra></extra>",
                ))
            if not proj_s.empty:
                fig_asia.add_trace(go.Scatter(
                    x=list(proj_s.index), y=list(proj_s.values / 1000),
                    line=dict(color=c, width=2, dash="dash"), showlegend=False, legendgroup=code,
                    hovertemplate=f"<b>{cname} (proj)</b><br>%{{x}}: $%{{y:.2f}}T<extra></extra>",
                ))

        fig_asia.add_vrect(x0=1980, x1=HIST_END, fillcolor="gray", opacity=0.05, line_width=0)
        fig_asia.add_vline(x=HIST_END, line_dash="dot", line_color="gray", line_width=1)
        fig_asia.update_layout(
            height=520, hovermode="x unified",
            xaxis_title="Year", yaxis_title="GDP (Trillions, 2021 USD PPP)",
            plot_bgcolor="white", paper_bgcolor="white",
            xaxis=dict(showgrid=True, gridcolor="#eee"),
            yaxis=dict(showgrid=True, gridcolor="#eee"),
            legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
            margin=dict(t=40, b=40),
        )
        st.plotly_chart(fig_asia, use_container_width=True)

        st.markdown("##### GDP Growth Rates — Key Asian Economies (%/yr)")
        fig_growth = go.Figure()
        for i, code in enumerate(asia_available[:6]):
            cname = name_map.get(code, code)
            yr_avail = sorted([y for y in range(1981, min(HIST_END + 1, 2031)) if y in gdp_df.columns])
            if code not in gdp_df.index or len(yr_avail) < 2:
                continue
            vals = []
            for yr in yr_avail:
                prev = yr - 1
                if prev in gdp_df.columns:
                    v_now = gdp_df.loc[code, yr]
                    v_prev = gdp_df.loc[code, prev]
                    if v_prev and v_prev > 0:
                        vals.append((yr, (np.log(v_now) - np.log(v_prev)) * 100))
            if vals:
                ys, gs = zip(*vals)
                fig_growth.add_trace(go.Bar(
                    x=list(ys), y=list(gs), name=cname,
                    hovertemplate=f"<b>{cname}</b><br>%{{x}}: %{{y:.1f}}%<extra></extra>",
                ))
        fig_growth.update_layout(
            height=400, barmode="group",
            xaxis_title="Year", yaxis_title="Annual GDP Growth (%)",
            plot_bgcolor="white", paper_bgcolor="white",
            xaxis=dict(showgrid=True, gridcolor="#eee"),
            yaxis=dict(showgrid=True, gridcolor="#eee", zeroline=True, zerolinecolor="black"),
            legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
            margin=dict(t=30, b=40),
        )
        st.plotly_chart(fig_growth, use_container_width=True)

    # ── Slide 25: Forecast comparison ────────────────────────────────────────
    with ppt3:
        st.markdown("##### 2030 GDP Forecasts — This Model vs External Benchmarks")
        st.caption(
            "Comparison of 2030 projections across major economies. "
            "'This Model' uses your current parameter settings."
        )
        # Hubbard & Sharma (2016, inflation-adjusted) and 2017 White Paper values from the Excel
        hs_2030 = {
            "CHN": 30396.04, "IND": 20722.36, "USA": 29076.50,
            "IDN": 5418.75,  "JPN": 5945.67,  "DEU": 7208.83,
            "AUS": 1827.71,
        }
        wp_2030 = {
            "CHN": 48964.80, "IND": 24621.60, "USA": 26920.00,
            "JPN": 6222.40,
        }

        compare_2030 = sorted(hs_2030.keys(), key=lambda c: -hs_2030.get(c, 0))
        model_2030 = {}
        for code in compare_2030:
            if code in gdp_df.index and 2030 in gdp_df.columns:
                model_2030[code] = float(gdp_df.loc[code, 2030])

        fig_cmp = go.Figure()
        x = [name_map.get(c, c) for c in compare_2030]
        fig_cmp.add_trace(go.Bar(
            name="This Model (Excel baseline)",
            x=x,
            y=[hs_2030.get(c, 0) / 1000 for c in compare_2030],
            marker_color="#2196F3",
        ))
        fig_cmp.add_trace(go.Bar(
            name="2017 White Paper",
            x=x,
            y=[wp_2030.get(c, 0) / 1000 for c in compare_2030],
            marker_color="#FF9800",
        ))
        fig_cmp.add_trace(go.Bar(
            name=f"Live Model (β={beta:.3f})",
            x=x,
            y=[model_2030.get(c, 0) / 1000 for c in compare_2030],
            marker_color="#4CAF50",
        ))
        fig_cmp.update_layout(
            height=460, barmode="group",
            xaxis_title="Country", yaxis_title="GDP 2030 (Trillions, 2021 USD PPP)",
            plot_bgcolor="white", paper_bgcolor="white",
            xaxis=dict(showgrid=False),
            yaxis=dict(showgrid=True, gridcolor="#eee"),
            legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
            margin=dict(t=40, b=40),
        )
        st.plotly_chart(fig_cmp, use_container_width=True)

    # ── China vs G7 vs North America ─────────────────────────────────────────
    with ppt4:
        st.markdown("##### China vs G7 vs North America — GDP (Trillions, 2021 USD PPP)")
        china_g7_yrs = [y for y in range(1980, proj_end + 1) if y in gdp_df.columns]
        g7_codes = [c for c in G7 if c in gdp_df.index]
        na_codes = [c for c in ["USA", "CAN", "MEX"] if c in gdp_df.index]

        china_s  = gdp_df.loc["CHN", china_g7_yrs].dropna() if "CHN" in gdp_df.index else pd.Series()
        india_s  = gdp_df.loc["IND", china_g7_yrs].dropna() if "IND" in gdp_df.index else pd.Series()
        g7_s     = gdp_df.loc[g7_codes, china_g7_yrs].sum()
        na_s     = gdp_df.loc[na_codes, china_g7_yrs].sum()
        chin_ind = (
            gdp_df.loc[["CHN", "IND"], china_g7_yrs].sum()
            if all(c in gdp_df.index for c in ["CHN", "IND"]) else pd.Series()
        )

        fig_g7 = go.Figure()
        for label_str, series, color in [
            ("China", china_s, "#E53935"),
            ("India", india_s, "#FB8C00"),
            ("China + India", chin_ind, "#D81B60"),
            ("G7", g7_s, "#1E88E5"),
            ("North America", na_s, "#43A047"),
        ]:
            if series.empty:
                continue
            hist_s2 = series[series.index <= HIST_END]
            proj_s2 = series[series.index > HIST_END]
            if not hist_s2.empty:
                fig_g7.add_trace(go.Scatter(
                    x=list(hist_s2.index), y=list(hist_s2.values / 1000),
                    name=label_str, line=dict(color=color, width=2.5),
                    hovertemplate=f"<b>{label_str}</b><br>%{{x}}: $%{{y:.1f}}T<extra></extra>",
                ))
            if not proj_s2.empty:
                fig_g7.add_trace(go.Scatter(
                    x=list(proj_s2.index), y=list(proj_s2.values / 1000),
                    line=dict(color=color, width=2.5, dash="dash"),
                    showlegend=False, legendgroup=label_str,
                    hovertemplate=f"<b>{label_str} (proj)</b><br>%{{x}}: $%{{y:.1f}}T<extra></extra>",
                ))

        # Crossover annotations
        fig_g7.add_vrect(x0=1980, x1=HIST_END, fillcolor="gray", opacity=0.05, line_width=0)
        fig_g7.add_vline(x=HIST_END, line_dash="dot", line_color="gray", line_width=1)
        fig_g7.update_layout(
            height=520, hovermode="x unified",
            xaxis_title="Year", yaxis_title="GDP (Trillions, 2021 USD PPP)",
            plot_bgcolor="white", paper_bgcolor="white",
            xaxis=dict(showgrid=True, gridcolor="#eee"),
            yaxis=dict(showgrid=True, gridcolor="#eee"),
            legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
            margin=dict(t=40, b=40),
        )
        st.plotly_chart(fig_g7, use_container_width=True)


# ════════════════════════════════════════════════════════════════════════════════
# TAB 6 — Technical
# ════════════════════════════════════════════════════════════════════════════════
with tab_tech:
    st.subheader("Technical Details")

    tech1, tech2, tech3 = st.tabs([
        "Kernel Regression Curve",
        "Comparator Countries",
        "Model Equations",
    ])

    with tech1:
        st.markdown("##### Gaussian Kernel Regression: GCI Score → Long-Run Steady State (LSS)")
        st.caption(
            "The kernel curve maps each country's Global Competitiveness Index (GCI) score "
            "to an estimated long-run labour productivity relative to the USA (natural log). "
            "Adjust the bandwidth slider in the sidebar to see how it affects the smoothing."
        )

        valid_comp = comp_df.dropna(subset=["gci", "log_prod"])
        gci_arr  = valid_comp["gci"].values.astype(float)
        prod_arr = valid_comp["log_prod"].values.astype(float)

        gci_grid = np.linspace(35, 90, 400)
        lss_curve = kernel_curve(gci_grid, gci_arr, prod_arr, bandwidth)

        fig_kern = go.Figure()
        # Scatter of comparator countries
        fig_kern.add_trace(go.Scatter(
            x=gci_arr, y=prod_arr, mode="markers",
            name="Comparator countries",
            marker=dict(size=8, color="#888", opacity=0.7),
            text=valid_comp["country"].values,
            hovertemplate="<b>%{text}</b><br>GCI: %{x:.1f}<br>log(rel. prod): %{y:.3f}<extra></extra>",
        ))
        # Kernel curve
        fig_kern.add_trace(go.Scatter(
            x=gci_grid, y=lss_curve,
            name=f"Kernel curve (BW={bandwidth:.2f})",
            line=dict(color="#E53935", width=2.5),
            hovertemplate="GCI: %{x:.1f} → LSS: %{y:.3f}<extra></extra>",
        ))
        # Highlight selected countries
        for code in selected:
            gci_v = kernel_df.loc[code, "gci"] if code in kernel_df.index else None
            lss_v = lss_map.get(code)
            if gci_v and lss_v and not np.isnan(gci_v):
                fig_kern.add_trace(go.Scatter(
                    x=[gci_v], y=[lss_v], mode="markers+text",
                    text=[code], textposition="top center",
                    marker=dict(size=12, color=COLORS[selected.index(code) % len(COLORS)], symbol="star"),
                    name=name_map.get(code, code), showlegend=True,
                    hovertemplate=f"<b>{name_map.get(code, code)}</b><br>GCI: {gci_v:.1f}<br>LSS: {lss_v:.3f}<extra></extra>",
                ))
        fig_kern.update_layout(
            height=480, xaxis_title="GCI Score",
            yaxis_title="ln(Labour Productivity / USA)",
            plot_bgcolor="white", paper_bgcolor="white",
            xaxis=dict(showgrid=True, gridcolor="#eee"),
            yaxis=dict(showgrid=True, gridcolor="#eee"),
            legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
            margin=dict(t=30, b=40),
        )
        st.plotly_chart(fig_kern, use_container_width=True)

    with tech2:
        st.markdown("##### Comparator Countries Used in Kernel Estimation")
        disp = comp_df[["country", "code", "gci", "log_prod", "comparator"]].copy()
        disp.columns = ["Country", "Code", "GCI Score", "ln(Rel. Productivity)", "Comparator"]
        disp["LSS (re-estimated)"] = disp["Code"].map(lss_map).round(4)
        disp["LSS as % of USA"] = (np.exp(disp["LSS (re-estimated)"]) * 100).round(1)
        st.dataframe(disp.set_index("Code").sort_values("GCI Score", ascending=False), use_container_width=True)

    with tech3:
        st.markdown("##### Model Equations")

        st.markdown("**1. Labour Productivity**")
        st.latex(r"\text{Productivity}_t = \frac{\text{GDP}_t}{\text{WAP}_t}")

        st.markdown("**2. Relative Productivity (vs USA)**")
        st.latex(r"y_t = \frac{\text{Productivity}_{\text{country},t}}{\text{Productivity}_{\text{USA},t}}")

        st.markdown("**3. Long-Run Steady State (Kernel Regression)**")
        st.latex(r"\ln(y^*) = \frac{\sum_i K\!\left(\frac{\text{GCI} - \text{GCI}_i}{h}\right) \ln(y_i)}{\sum_i K\!\left(\frac{\text{GCI} - \text{GCI}_i}{h}\right)}")
        st.caption("where K(u) = exp(-u²/2) is the Gaussian kernel and h is the bandwidth.")

        st.markdown("**4. Conditional Convergence (projection)**")
        st.latex(r"\ln(y_{t+1}) = \ln(y_t) + \beta \cdot [\ln(y^*) - \ln(y_t)]")

        st.markdown("**5. GDP per Capita Projection**")
        st.latex(r"\text{GDPPC}_t = \text{GDPPC}_{2024} \cdot \frac{y_t}{y_{2024}} \cdot e^{g_{\text{USA}} \cdot (t - 2024)} \cdot \frac{\text{WAP}_t / \text{Pop}_t}{\text{WAP}_{2024} / \text{Pop}_{2024}}")

        st.markdown("**6. Total GDP**")
        st.latex(r"\text{GDP}_t = \text{GDPPC}_t \times \text{Population}_t / 10^6 \quad \text{(billions USD PPP)}")

        st.markdown("**Parameters**")
        st.dataframe(pd.DataFrame({
            "Symbol":      ["β", "g_USA", "h"],
            "Description": ["Annual convergence speed", "US productivity growth", "Kernel bandwidth"],
            "Default":     ["0.025", "1.8%/yr", "7.76"],
            "Current":     [f"{beta:.3f}", f"{g_usa*100:.1f}%/yr", f"{bandwidth:.2f}"],
        }), use_container_width=True, hide_index=True)


# ════════════════════════════════════════════════════════════════════════════════
# DOWNLOAD SECTION
# ════════════════════════════════════════════════════════════════════════════════
st.divider()
st.subheader("Export Forecast Data to Excel")

with st.expander("Configure and download", expanded=True):
    dl_col1, dl_col2, dl_col3 = st.columns(3)

    with dl_col1:
        dl_year_start = st.number_input("From year", min_value=1980, max_value=proj_end, value=2025, step=1)
        dl_year_end   = st.number_input("To year",   min_value=1980, max_value=proj_end, value=proj_end, step=1)

    with dl_col2:
        dl_countries = st.multiselect(
            "Countries to export",
            options=all_codes,
            default=selected,
            format_func=label,
        )
        dl_select_all = st.checkbox("Select all 140 countries")
        if dl_select_all:
            dl_countries = all_codes

    with dl_col3:
        dl_metrics = st.multiselect(
            "Metrics to include",
            ["GDP per Capita (USD)", "GDP Total (Billions)", "GDP Total (Trillions)",
             "Population (Thousands)", "WAP (Thousands)", "Rel. Productivity vs USA"],
            default=["GDP per Capita (USD)", "GDP Total (Billions)"],
        )
        dl_include_hist = st.checkbox("Include historical data (pre-2025)", value=False)

    def build_excel_export(
        countries, year_start, year_end, metrics, include_hist
    ):
        """Build a multi-sheet Excel workbook as bytes."""
        import io
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            return None

        wb = Workbook()
        wb.remove(wb.active)  # remove default sheet

        years = list(range(year_start, year_end + 1))
        proj_years_only = [y for y in years if y > HIST_END]
        hist_years_only = [y for y in years if y <= HIST_END]

        header_font  = Font(bold=True, color="FFFFFF")
        header_fill  = PatternFill("solid", fgColor="1F3864")
        proj_fill    = PatternFill("solid", fgColor="E8F4FD")
        hist_fill    = PatternFill("solid", fgColor="F2F2F2")
        center_align = Alignment(horizontal="center")

        def write_metric_sheet(ws_name, get_hist_val, get_proj_val, fmt=",.0f"):
            ws = wb.create_sheet(ws_name)
            # Header row
            ws.cell(1, 1, "Country").font = Font(bold=True)
            ws.cell(1, 2, "Code").font = Font(bold=True)
            for ci, yr in enumerate(years, start=3):
                cell = ws.cell(1, ci, str(yr))
                cell.font = header_font
                cell.fill = proj_fill if yr > HIST_END else hist_fill
                cell.alignment = center_align
            ws.cell(1, 1).fill = PatternFill("solid", fgColor="2E4057")
            ws.cell(1, 1).font = Font(bold=True, color="FFFFFF")
            ws.cell(1, 2).fill = PatternFill("solid", fgColor="2E4057")
            ws.cell(1, 2).font = Font(bold=True, color="FFFFFF")

            # Data rows
            for ri, code in enumerate(countries, start=2):
                cname = name_map.get(code, code)
                ws.cell(ri, 1, cname)
                ws.cell(ri, 2, code)
                for ci, yr in enumerate(years, start=3):
                    if yr <= HIST_END:
                        if not include_hist:
                            continue
                        v = get_hist_val(code, yr)
                    else:
                        v = get_proj_val(code, yr)
                    if v is not None and not (isinstance(v, float) and np.isnan(v)):
                        ws.cell(ri, ci, round(float(v), 2))

            # Column widths
            ws.column_dimensions["A"].width = 22
            ws.column_dimensions["B"].width = 7
            for ci in range(3, len(years) + 3):
                ws.column_dimensions[get_column_letter(ci)].width = 11

            # Freeze header
            ws.freeze_panes = "C2"

        # Pre-compute projections for all download countries
        dl_lss = compute_lss(bandwidth)
        dl_projs = {}
        for code in countries:
            lss_v = dl_lss.get(code)
            if lss_v is None or not np.isfinite(lss_v):
                continue
            proj = project_country(
                code, gdppc_df, pop_df, wap_df, rel_prod_df,
                lss_v, beta, g_usa, base_year=HIST_END, end_year=year_end,
            )
            if proj is not None:
                dl_projs[code] = proj

        def safe_proj(code, yr, col):
            if code in dl_projs and yr in dl_projs[code].index:
                return dl_projs[code].loc[yr, col]
            return None

        # Write requested sheets
        if "GDP per Capita (USD)" in metrics:
            write_metric_sheet(
                "GDP per Capita (USD PPP)",
                get_hist_val=lambda c, y: float(gdppc_df.loc[c, y]) if c in gdppc_df.index and y in gdppc_df.columns else None,
                get_proj_val=lambda c, y: safe_proj(c, y, "gdppc"),
            )

        if "GDP Total (Billions)" in metrics:
            write_metric_sheet(
                "GDP Total (Billions)",
                get_hist_val=lambda c, y: float(gdp_df.loc[c, y]) if c in gdp_df.index and y in gdp_df.columns else None,
                get_proj_val=lambda c, y: safe_proj(c, y, "gdp_bn"),
            )

        if "GDP Total (Trillions)" in metrics:
            write_metric_sheet(
                "GDP Total (Trillions)",
                get_hist_val=lambda c, y: float(gdp_df.loc[c, y]) / 1000 if c in gdp_df.index and y in gdp_df.columns else None,
                get_proj_val=lambda c, y: safe_proj(c, y, "gdp_bn") / 1000 if safe_proj(c, y, "gdp_bn") is not None else None,
                fmt=",.3f",
            )

        if "Population (Thousands)" in metrics:
            write_metric_sheet(
                "Population (Thousands)",
                get_hist_val=lambda c, y: float(pop_df.loc[c, y]) if c in pop_df.index and y in pop_df.columns else None,
                get_proj_val=lambda c, y: float(pop_df.loc[c, y]) if c in pop_df.index and y in pop_df.columns else None,
            )

        if "WAP (Thousands)" in metrics:
            write_metric_sheet(
                "WAP (Thousands)",
                get_hist_val=lambda c, y: float(wap_df.loc[c, y]) if c in wap_df.index and y in wap_df.columns else None,
                get_proj_val=lambda c, y: float(wap_df.loc[c, y]) if c in wap_df.index and y in wap_df.columns else None,
            )

        if "Rel. Productivity vs USA" in metrics:
            def _rel_prod_proj(c, y):
                c_gdppc = safe_proj(c, y, "gdppc")
                if c_gdppc is None:
                    return None
                usa_gdppc = safe_proj("USA", y, "gdppc")
                if usa_gdppc is None and "USA" in gdppc_df.index and y in gdppc_df.columns:
                    usa_gdppc = float(gdppc_df.loc["USA", y])
                if not usa_gdppc:
                    return None
                # Approximate relative productivity as GDPPC ratio × (WAP/Pop adjustment)
                c_wap = float(wap_df.loc[c, y]) if c in wap_df.index and y in wap_df.columns else None
                c_pop = float(pop_df.loc[c, y]) if c in pop_df.index and y in pop_df.columns else None
                u_wap = float(wap_df.loc["USA", y]) if "USA" in wap_df.index and y in wap_df.columns else None
                u_pop = float(pop_df.loc["USA", y]) if "USA" in pop_df.index and y in pop_df.columns else None
                if all(v and v > 0 for v in [c_wap, c_pop, u_wap, u_pop]):
                    return (c_gdppc / usa_gdppc) * ((u_wap / u_pop) / (c_wap / c_pop))
                return c_gdppc / usa_gdppc

            write_metric_sheet(
                "Rel. Productivity vs USA",
                get_hist_val=lambda c, y: float(rel_prod_df.loc[c, y]) if c in rel_prod_df.index and y in rel_prod_df.columns else None,
                get_proj_val=_rel_prod_proj,
                fmt=",.4f",
            )

        # Add a metadata sheet
        ws_meta = wb.create_sheet("Model Parameters", 0)
        ws_meta.column_dimensions["A"].width = 30
        ws_meta.column_dimensions["B"].width = 20
        meta_rows = [
            ("CCM GDP Projection Dashboard", ""),
            ("", ""),
            ("Export Date", pd.Timestamp.now().strftime("%Y-%m-%d")),
            ("Year Range", f"{year_start} – {year_end}"),
            ("Countries", len(countries)),
            ("", ""),
            ("Model Parameters", ""),
            ("β (Convergence Speed)", beta),
            ("g_USA (US Productivity Growth)", f"{g_usa*100:.2f}%/yr"),
            ("Kernel Bandwidth", bandwidth),
            ("GCI Weight A (Basic Req.)", w_a),
            ("GCI Weight B (Efficiency)", w_b),
            ("GCI Weight C (Innovation)", w_c),
            ("", ""),
            ("Notes", "Projections use 2021 USD PPP. Historical data: World Bank WDI."),
            ("", "Shaded columns = projected values."),
        ]
        for i, (k, v) in enumerate(meta_rows, start=1):
            ws_meta.cell(i, 1, k)
            ws_meta.cell(i, 2, v if not isinstance(v, float) else round(v, 6))
            if k and not v == "":
                ws_meta.cell(i, 1).font = Font(bold=True)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    if not dl_countries:
        st.warning("Select at least one country to export.")
    elif dl_year_start > dl_year_end:
        st.warning("Start year must be ≤ end year.")
    else:
        n_years  = dl_year_end - dl_year_start + 1
        n_sheets = len(dl_metrics)
        st.caption(
            f"Export: **{len(dl_countries)} countries** × **{n_years} years** "
            f"({dl_year_start}–{dl_year_end}) × **{n_sheets} metric sheet(s)**"
        )
        if st.button("📥 Generate & Download Excel", type="primary"):
            with st.spinner("Building Excel file…"):
                xlsx_bytes = build_excel_export(
                    dl_countries, dl_year_start, dl_year_end,
                    dl_metrics, dl_include_hist,
                )
            if xlsx_bytes:
                fname = (
                    f"CCM_GDP_Forecast_{dl_year_start}-{dl_year_end}"
                    f"_{len(dl_countries)}countries.xlsx"
                )
                st.download_button(
                    label="💾 Click here to save the file",
                    data=xlsx_bytes,
                    file_name=fname,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
                st.success(f"Ready: **{fname}** ({len(dl_countries)} countries, {n_years} years, {n_sheets} sheets)")
            else:
                st.error("Failed to generate Excel file.")


# ─── Footer ──────────────────────────────────────────────────────────────────
st.divider()
st.caption(
    "Data: World Development Indicators (World Bank) · GCI: World Economic Forum · "
    "Model: Conditional Convergence (Solow-Swan framework, Kernel-estimated steady states) · "
    "Projections from 2025 onward. Historical shading = 1980–2024."
)
